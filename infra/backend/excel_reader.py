from __future__ import annotations

import io
import os
import re
import unicodedata
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

MAX_SCAN_ROWS = int(os.environ.get("MAX_SCAN_ROWS", "20000"))
MAX_CELL_CHARS = 2000
VEHICLES = ("kamyon", "kirkayak", "tir")


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)[:MAX_CELL_CHARS]


def _normalize(value: Any) -> str:
    text = _text(value).replace("ı", "i").casefold()
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(char)
    )


def search_workbook(
    workbook_bytes: bytes, question: str, source: str, limit: int
) -> tuple[list[dict[str, Any]], bool]:
    terms = [_normalize(term) for term in question.split() if len(term) > 1]
    if not terms:
        raise ValueError("Soru en az bir anlamlı arama kelimesi içermeli.")

    workbook = load_workbook(
        io.BytesIO(workbook_bytes), read_only=True, data_only=True
    )
    matches: list[dict[str, Any]] = []
    scanned = 0

    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            transport_matches = _transport_prices(rows, sheet.title, question, source)
            if transport_matches:
                matches.extend(transport_matches)
                continue

            header_index, header_row = max(
                enumerate(rows[:10]),
                key=lambda item: sum(value is not None for value in item[1]),
                default=(0, ()),
            )
            headers = _headers(header_row)

            for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
                scanned += 1
                if scanned > MAX_SCAN_ROWS:
                    return _top(matches, limit), True

                values = [_text(value) for value in row[: len(headers)]]
                searchable = _normalize(" ".join([sheet.title, *headers, *values]))
                score = sum(term in searchable for term in terms)
                if score:
                    matches.append(
                        {
                            "source": source,
                            "sheet": sheet.title,
                            "row": row_number,
                            "score": score,
                            "values": {
                                header: value
                                for header, value in zip(headers, values)
                                if value
                            },
                        }
                    )
    finally:
        workbook.close()

    # ponytail: lexical row ranking; replace with schema-specific filters when
    # workbook column contracts are stable.
    return _top(matches, limit), False


def _transport_prices(
    rows: list[tuple[Any, ...]], sheet: str, question: str, source: str
) -> list[dict[str, Any]]:
    """Return the exact destination row for FTL price questions."""
    if "cikis" not in _normalize(sheet) or len(rows) < 4:
        return []

    normalized_question = f" {re.sub(r'\W+', ' ', _normalize(question)).strip()} "
    candidates = []
    for row_number, row in enumerate(rows[3:], start=4):
        province = _text(row[1] if len(row) > 1 else None).strip()
        district = _text(row[2] if len(row) > 2 else None).strip()
        if not province or not district:
            continue
        province_key, district_key = _normalize(province), _normalize(district)
        if f" {district_key} " in normalized_question:
            candidates.append((2, len(district_key), row_number, row, province, district))
        elif f" {province_key} " in normalized_question:
            center = district_key in {province_key, "merkez"}
            candidates.append((1 if center else 0, len(province_key), row_number, row, province, district))

    if not candidates:
        return []
    best = max(candidates, key=lambda item: (item[0], item[1], -item[2]))
    _, _, row_number, row, province, district = best
    values = {
        "Çıkış": sheet.removesuffix(" Çıkış"),
        "İl": province,
        "İlçe": district,
        "Kamyon": _money(row[3] if len(row) > 3 else None),
        "Kırkayak": _money(row[4] if len(row) > 4 else None),
        "Tır": _money(row[5] if len(row) > 5 else None),
        "Kapsam": "FTL/komple taşıma",
    }
    requested = [vehicle for vehicle in VEHICLES if f" {vehicle} " in normalized_question]
    score = 3 + bool(requested)
    return [{"source": source, "sheet": sheet, "row": row_number, "score": score, "values": values}]


def _money(value: Any) -> str:
    if isinstance(value, (int, float)):
        return f"{value:.0f} TL"
    return _text(value)


def _headers(row: tuple[Any, ...]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(row[:50], start=1):
        base = _text(value).strip() or f"column_{index}"
        seen[base] = seen.get(base, 0) + 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return headers


def _top(matches: list[dict[str, Any]], limit: int) -> list[dict[str, Any]]:
    return sorted(matches, key=lambda item: (-item["score"], item["row"]))[:limit]
