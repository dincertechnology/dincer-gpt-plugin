import io
import unittest

from openpyxl import Workbook

from excel_reader import search_workbook


class ExcelReaderTest(unittest.TestCase):
    def test_finds_turkish_text_and_limits_output(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Fiyatlar"
        sheet.append(["Çıkış", "Varış", "Fiyat"])
        sheet.append(["İstanbul", "Ankara", 1250])
        sheet.append(["İzmir", "Bursa", 900])
        buffer = io.BytesIO()
        workbook.save(buffer)

        matches, truncated = search_workbook(
            buffer.getvalue(), "istanbul ankara", "tasima", 1
        )

        self.assertFalse(truncated)
        self.assertEqual(1, len(matches))
        self.assertEqual("1250", matches[0]["values"]["Fiyat"])

    def test_ftl_province_uses_named_center_and_returns_both_origins(self):
        workbook = Workbook()
        for index, (title, price) in enumerate(
            (("Gebze Çıkış", 89306.83), ("İzmir Çıkış", 83297.11))
        ):
            sheet = workbook.active if index == 0 else workbook.create_sheet()
            sheet.title = title
            sheet.append([])
            sheet.append([None, "81 İl Fiyat Teklifi"])
            sheet.append([None, "İl", "İlçe", "Kamyon", "Kırkayak", "Tır"])
            sheet.append([None, "Iğdır", "Iğdır", price, 90000, 100000])
            sheet.append([None, "Iğdır", "Aralık", 99999, 99999, 99999])
        buffer = io.BytesIO()
        workbook.save(buffer)

        matches, truncated = search_workbook(
            buffer.getvalue(), "Iğdır'da kamyon fiyatı nedir?", "tasima", 8
        )

        self.assertFalse(truncated)
        self.assertEqual(["89307 TL", "83297 TL"], [m["values"]["Kamyon"] for m in matches])
        self.assertTrue(all(m["values"]["İlçe"] == "Iğdır" for m in matches))

    def test_ftl_district_overrides_province_center(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Gebze Çıkış"
        sheet.append([])
        sheet.append([None, "81 İl Fiyat Teklifi"])
        sheet.append([None, "İl", "İlçe", "Kamyon", "Kırkayak", "Tır"])
        sheet.append([None, "Iğdır", "Iğdır", 100, 200, 300])
        sheet.append([None, "Iğdır", "Aralık", 400, 500, 600])
        buffer = io.BytesIO()
        workbook.save(buffer)

        matches, _ = search_workbook(buffer.getvalue(), "Iğdır Aralık tır", "tasima", 8)

        self.assertEqual("Aralık", matches[0]["values"]["İlçe"])
        self.assertEqual("600 TL", matches[0]["values"]["Tır"])


if __name__ == "__main__":
    unittest.main()
